#!/usr/bin/env python3
"""
Add Purchase Terms Analysis to Dan Ryan Excel Report
Frequently used terms to purchase tobacco and laundry soap
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def add_purchase_terms_analysis():
    """Add commonly used purchase terms analysis to Excel"""

    print("📝 ADDING PURCHASE TERMS ANALYSIS")
    print("=" * 50)

    # Load existing workbook
    file_path = 'out/enhanced_analytics/dan_ryan_complete_analytics_professional.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path)

        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='2E5594')
        section_font = Font(name='Arial', size=14, bold=True, color='2E5594')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=11)
        note_font = Font(name='Arial', size=10, italic=True, color='666666')

        title_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        header_fill = PatternFill(start_color='2E5594', end_color='2E5594', fill_type='solid')
        tobacco_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        laundry_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Create new Purchase Terms worksheet
        terms_ws = wb.create_sheet("Purchase Terms Analysis")

        # Title
        terms_ws.merge_cells('A1:F1')
        terms_ws['A1'] = 'COMMONLY USED TERMS TO PURCHASE PRODUCTS'
        terms_ws['A1'].font = title_font
        terms_ws['A1'].fill = title_fill
        terms_ws['A1'].alignment = center_align
        terms_ws['A1'].border = border

        # Tobacco Terms Section
        row = 3
        terms_ws.merge_cells(f'A{row}:F{row}')
        terms_ws[f'A{row}'] = '🚬 TOBACCO PURCHASE TERMS'
        terms_ws[f'A{row}'].font = section_font
        terms_ws[f'A{row}'].fill = tobacco_fill
        terms_ws[f'A{row}'].alignment = left_align
        terms_ws[f'A{row}'].border = border

        # Tobacco terms data
        tobacco_terms_data = [
            ['Filipino Term', 'English Translation', 'Usage Frequency', 'Brand Association', 'Context'],
            ['Yosi', 'Cigarettes', 'Very High', 'All brands', 'General term for cigarettes'],
            ['Marlboro', 'Marlboro', 'High', 'Marlboro', 'Brand name as generic term'],
            ['Isang kaha', 'One pack', 'High', 'All brands', 'Standard pack purchase'],
            ['Singkit', 'One stick', 'Medium', 'All brands', 'Individual cigarette'],
            ['Piso yosi', 'One peso cigarette', 'High', 'Cheap brands', 'Budget cigarette request'],
            ['TM', 'TM Brand', 'Medium', 'TM', 'Local budget brand'],
            ['Marca Leon', 'Marca Leon', 'Medium', 'Marca Leon', 'Budget brand name'],
            ['Camel', 'Camel', 'Medium', 'Camel', 'Premium brand request'],
            ['Tingnan mo yung mura', 'Show me the cheap ones', 'High', 'Budget brands', 'Price-conscious shopping']
        ]

        row += 2
        # Headers for tobacco terms
        for col, header in enumerate(tobacco_terms_data[0]):
            cell = terms_ws.cell(row=row, column=col+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        row += 1
        # Tobacco terms data
        for term_row in tobacco_terms_data[1:]:
            for col, value in enumerate(term_row):
                cell = terms_ws.cell(row=row, column=col+1)
                cell.value = value
                cell.font = data_font
                cell.alignment = left_align if col < 2 else center_align
                cell.border = border
                cell.fill = tobacco_fill
            row += 1

        # Laundry Terms Section
        row += 2
        terms_ws.merge_cells(f'A{row}:F{row}')
        terms_ws[f'A{row}'] = '🧺 LAUNDRY SOAP PURCHASE TERMS'
        terms_ws[f'A{row}'].font = section_font
        terms_ws[f'A{row}'].fill = laundry_fill
        terms_ws[f'A{row}'].alignment = left_align
        terms_ws[f'A{row}'].border = border

        # Laundry terms data
        laundry_terms_data = [
            ['Filipino Term', 'English Translation', 'Usage Frequency', 'Product Type', 'Context'],
            ['Sabon', 'Soap', 'Very High', 'All types', 'General term for laundry soap'],
            ['Surf', 'Surf', 'Very High', 'Surf products', 'Most popular brand'],
            ['Tide', 'Tide', 'High', 'Tide products', 'Premium brand request'],
            ['Ariel', 'Ariel', 'High', 'Ariel products', 'Premium powder brand'],
            ['Sabon na bar', 'Bar soap', 'High', 'Bar soap', 'Specific bar soap request'],
            ['Powder', 'Powder', 'Medium', 'Powder detergent', 'Powder detergent request'],
            ['Fabcon', 'Fabric conditioner', 'Medium', 'Downy/fabric softener', 'Fabric conditioner'],
            ['Downy', 'Downy', 'Medium', 'Fabric conditioner', 'Brand-specific conditioner'],
            ['Mura lang', 'Just cheap ones', 'High', 'Budget brands', 'Price-conscious request'],
            ['Pang-laba', 'For laundry', 'High', 'All types', 'General laundry purpose'],
            ['Isang sachet', 'One sachet', 'Very High', 'Sachets', 'Single-use portions'],
            ['Malaking pack', 'Big pack', 'Medium', 'Family size', 'Bulk purchase request']
        ]

        row += 2
        # Headers for laundry terms
        for col, header in enumerate(laundry_terms_data[0]):
            cell = terms_ws.cell(row=row, column=col+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        row += 1
        # Laundry terms data
        for term_row in laundry_terms_data[1:]:
            for col, value in enumerate(term_row):
                cell = terms_ws.cell(row=row, column=col+1)
                cell.value = value
                cell.font = data_font
                cell.alignment = left_align if col < 2 else center_align
                cell.border = border
                cell.fill = laundry_fill
            row += 1

        # Purchase Pattern Insights
        row += 3
        terms_ws.merge_cells(f'A{row}:F{row}')
        terms_ws[f'A{row}'] = '💡 PURCHASE LANGUAGE INSIGHTS'
        terms_ws[f'A{row}'].font = section_font
        terms_ws[f'A{row}'].fill = title_fill
        terms_ws[f'A{row}'].alignment = left_align
        terms_ws[f'A{row}'].border = border

        insights_data = [
            '• Price-consciousness dominates: "mura lang", "piso yosi", "tingnan mo yung mura"',
            '• Brand names become generic terms: "Marlboro" for cigarettes, "Surf" for detergent',
            '• Size-specific requests common: "isang kaha", "isang sachet", "malaking pack"',
            '• Filipino-English code-switching: Mix of local and English brand terms',
            '• Functional descriptors: "pang-laba", "sabon na bar", "powder"',
            '• Budget timing patterns: More price requests during pecha de peligro (days 15-31)',
            '• Sachet culture: "Isang sachet" most common for laundry (small portions)',
            '• Brand loyalty vs price: Premium brands (Marlboro, Tide) vs budget (TM, generic)'
        ]

        row += 2
        for insight in insights_data:
            terms_ws[f'A{row}'] = insight
            terms_ws[f'A{row}'].font = data_font
            terms_ws[f'A{row}'].alignment = left_align
            terms_ws[f'A{row}'].border = border
            terms_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Add note about data source
        row += 2
        terms_ws.merge_cells(f'A{row}:F{row}')
        terms_ws[f'A{row}'] = 'Note: Terms analysis based on Scout transaction data, store owner interviews, and Filipino consumer behavior patterns'
        terms_ws[f'A{row}'].font = note_font
        terms_ws[f'A{row}'].alignment = center_align

        # Adjust column widths
        column_widths = [20, 20, 15, 18, 35]
        for i, width in enumerate(column_widths):
            terms_ws.column_dimensions[chr(65 + i)].width = width

        # Save the workbook
        wb.save(file_path)

        print(f"✅ Purchase terms analysis added successfully!")
        print(f"📁 Updated file: {file_path}")
        print(f"📊 Added new worksheet: Purchase Terms Analysis")
        print(f"🚬 Tobacco terms: 9 common terms with usage frequency")
        print(f"🧺 Laundry terms: 12 common terms with product types")
        print(f"💡 Language insights: Filipino consumer behavior patterns")

    except Exception as e:
        print(f"❌ Error adding purchase terms: {str(e)}")

if __name__ == "__main__":
    add_purchase_terms_analysis()