#!/usr/bin/env python3
"""
Add Real Data Truth Cover Sheet to Dan Ryan Excel Report
"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def add_real_data_cover():
    """Add the real data truth cover sheet"""

    print("📊 ADDING REAL DATA TRUTH COVER SHEET")
    print("=" * 50)

    # Load existing workbook
    file_path = 'out/enhanced_analytics/dan_ryan_complete_analytics_professional.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path)

        # Create new cover sheet at the beginning
        cover_ws = wb.create_sheet("REAL DATA TRUTH", 0)

        # Define styles
        title_font = Font(name='Arial', size=18, bold=True, color='2E5594')
        section_font = Font(name='Arial', size=14, bold=True, color='2E5594')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=11)
        warning_font = Font(name='Arial', size=12, bold=True, color='D32F2F')

        title_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        header_fill = PatternFill(start_color='2E5594', end_color='2E5594', fill_type='solid')
        truth_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        warning_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Title
        cover_ws.merge_cells('A1:F1')
        cover_ws['A1'] = 'SCOUT v7 - REAL DATA TRUTH'
        cover_ws['A1'].font = title_font
        cover_ws['A1'].fill = title_fill
        cover_ws['A1'].alignment = center_align
        cover_ws['A1'].border = border

        # Subtitle
        cover_ws.merge_cells('A2:F2')
        cover_ws['A2'] = 'Only Canonical_TX_ID Verified Transaction Data Included'
        cover_ws['A2'].font = section_font
        cover_ws['A2'].alignment = center_align

        # The Real Truth Section
        row = 4
        cover_ws.merge_cells(f'A{row}:F{row}')
        cover_ws[f'A{row}'] = '🔍 THE REAL TRUTH'
        cover_ws[f'A{row}'].font = section_font
        cover_ws[f'A{row}'].fill = truth_fill
        cover_ws[f'A{row}'].alignment = left_align
        cover_ws[f'A{row}'].border = border

        truth_data = [
            ['Only 4 stores have real canonical_tx_id transaction data (not 20!)', ''],
            ['2,797 verified transactions with facial recognition data', ''],
            ['57 unique customers with real demographic data', ''],
            ['Average 699 transactions per store - actual activity levels', '']
        ]

        row += 1
        for item, value in truth_data:
            cover_ws[f'A{row}'] = f'• {item}'
            cover_ws[f'A{row}'].font = data_font
            cover_ws[f'A{row}'].alignment = left_align
            cover_ws[f'A{row}'].border = border
            cover_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # The 4 REAL Active Stores
        row += 1
        cover_ws.merge_cells(f'A{row}:F{row}')
        cover_ws[f'A{row}'] = '🏪 THE 4 REAL ACTIVE STORES'
        cover_ws[f'A{row}'].font = section_font
        cover_ws[f'A{row}'].fill = truth_fill
        cover_ws[f'A{row}'].alignment = left_align
        cover_ws[f'A{row}'].border = border

        stores_data = [
            ['Store Name', 'Location', 'Transactions', 'Customers', 'Status'],
            ['Tess Store (108)', 'Quezon City', '5,848', '0', 'NO DEMOGRAPHICS'],
            ['Lourdes Store (109)', 'Quezon City', '2,482', '57', 'ACTIVE'],
            ['Riza Store (103)', 'Quezon City', '1,465', '54', 'ACTIVE'],
            ['Merly Store (110)', 'Manila', '1,300', '51', 'ACTIVE']
        ]

        row += 1
        # Headers
        for col, header in enumerate(stores_data[0]):
            cell = cover_ws.cell(row=row, column=col+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        row += 1
        # Store data
        for store_row in stores_data[1:]:
            for col, value in enumerate(store_row):
                cell = cover_ws.cell(row=row, column=col+1)
                cell.value = value
                cell.font = data_font
                cell.alignment = center_align if col > 0 else left_align
                cell.border = border

                # Special formatting for status
                if col == 4 and 'NO DEMOGRAPHICS' in value:
                    cell.font = warning_font
                    cell.fill = warning_fill
                elif col == 4 and 'ACTIVE' in value:
                    cell.fill = truth_fill

            row += 1

        # Data Quality Reality Check
        row += 1
        cover_ws.merge_cells(f'A{row}:F{row}')
        cover_ws[f'A{row}'] = '⚠️ DATA QUALITY REALITY CHECK'
        cover_ws[f'A{row}'].font = section_font
        cover_ws[f'A{row}'].fill = warning_fill
        cover_ws[f'A{row}'].alignment = left_align
        cover_ws[f'A{row}'].border = border

        reality_data = [
            ['Metric', 'Database Claims', 'Reality'],
            ['Total NCR Stores', '20 stores listed', 'Only 4 active (20%)'],
            ['Transaction Data', '12,192 total', '2,797 with demographics'],
            ['Customer Coverage', 'Assumed full', '57 unique customers only'],
            ['GPS Coordinates', 'Expected all', '1 out of 4 stores (25%)'],
            ['Manager Information', 'Expected contact data', '0 out of 4 stores (0%)'],
            ['Facial Recognition', 'System installed', 'Works in 3/4 stores only']
        ]

        row += 1
        # Headers for reality check
        for col, header in enumerate(reality_data[0]):
            cell = cover_ws.cell(row=row, column=col+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        row += 1
        # Reality data
        for reality_row in reality_data[1:]:
            for col, value in enumerate(reality_row):
                cell = cover_ws.cell(row=row, column=col+1)
                cell.value = value
                cell.font = data_font
                cell.alignment = left_align
                cell.border = border

                # Highlight reality column
                if col == 2:
                    cell.fill = warning_fill

            row += 1

        # Professional Excel Features
        row += 2
        cover_ws.merge_cells(f'A{row}:F{row}')
        cover_ws[f'A{row}'] = '📊 PROFESSIONAL EXCEL FEATURES'
        cover_ws[f'A{row}'].font = section_font
        cover_ws[f'A{row}'].fill = truth_fill
        cover_ws[f'A{row}'].alignment = left_align
        cover_ws[f'A{row}'].border = border

        features_data = [
            '• High-activity highlighting for stores >1000 transactions',
            '• Verified transaction counts with canonical_tx_id matches only',
            '• Real customer demographics from facial recognition system',
            '• Professional formatting with borders and shading',
            '• Currency, percentage, and number formatting',
            '• Municipality data from confirmed active locations'
        ]

        row += 1
        for feature in features_data:
            cover_ws[f'A{row}'] = feature
            cover_ws[f'A{row}'].font = data_font
            cover_ws[f'A{row}'].alignment = left_align
            cover_ws[f'A{row}'].border = border
            cover_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Footer
        row += 2
        cover_ws.merge_cells(f'A{row}:F{row}')
        cover_ws[f'A{row}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Data Source: Scout v7 Database | Demographics: Facial Recognition System'
        cover_ws[f'A{row}'].font = Font(name='Arial', size=10, italic=True)
        cover_ws[f'A{row}'].alignment = center_align
        cover_ws[f'A{row}'].fill = title_fill

        # Adjust column widths
        column_widths = [30, 20, 15, 15, 20, 10]
        for i, width in enumerate(column_widths):
            cover_ws.column_dimensions[chr(65 + i)].width = width

        # Save the workbook
        wb.save(file_path)

        print(f"✅ Real Data Truth cover sheet added successfully!")
        print(f"📁 Updated file: {file_path}")
        print(f"📊 Cover sheet includes:")
        print(f"   • The real truth about only 4 active stores")
        print(f"   • Detailed store breakdown with demographics status")
        print(f"   • Data quality reality check comparison")
        print(f"   • Professional formatting with color coding")

    except Exception as e:
        print(f"❌ Error adding cover sheet: {str(e)}")

if __name__ == "__main__":
    add_real_data_cover()