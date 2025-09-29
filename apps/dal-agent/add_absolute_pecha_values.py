#!/usr/bin/env python3
"""
Add Absolute Values for Pecha de Peligro Analysis
Show actual transaction counts alongside percentages
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def add_absolute_pecha_values():
    """Add absolute values to the Pecha de Peligro analysis"""

    print("📊 ADDING ABSOLUTE VALUES TO PECHA DE PELIGRO ANALYSIS")
    print("=" * 60)

    # Load existing workbook
    file_path = 'out/enhanced_analytics/dan_ryan_complete_analytics_professional.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path)

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=11)
        section_font = Font(name='Arial', size=14, bold=True, color='2E5594')

        header_fill = PatternFill(start_color='2E5594', end_color='2E5594', fill_type='solid')
        pecha_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Update Tobacco Analytics sheet
        tobacco_ws = wb['Tobacco Analytics']

        # Calculate absolute values for tobacco
        tobacco_pecha_data = [
            ['Brand', 'Total Purchases', 'Days 1-14', 'Days 15-31 (Pecha)', 'Pecha %', 'Pecha Count'],
            ['Marlboro', 108, 70, 38, '35.2%', 38],
            ['Camel', 52, 34, 18, '34.6%', 18],
            ['TM', 21, 13, 8, '38.1%', 8],
            ['Marca Leon', 11, 4, 7, '63.6%', 7],
            ['Winston', 2, 1, 1, '50.0%', 1],
            ['Chesterfield', 1, 1, 0, '0.0%', 0]
        ]

        # Find where to insert the new table (after existing purchase profile)
        # Insert around row 20-25 area
        start_row = 25

        tobacco_ws.merge_cells(f'A{start_row}:F{start_row}')
        tobacco_ws[f'A{start_row}'] = 'PECHA DE PELIGRO - ABSOLUTE VALUES & PERCENTAGES'
        tobacco_ws[f'A{start_row}'].font = section_font
        tobacco_ws[f'A{start_row}'].fill = pecha_fill
        tobacco_ws[f'A{start_row}'].alignment = left_align
        tobacco_ws[f'A{start_row}'].border = border

        # Add headers
        start_row += 2
        for col, header in enumerate(tobacco_pecha_data[0]):
            cell = tobacco_ws.cell(row=start_row, column=col+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Add data
        start_row += 1
        for row_data in tobacco_pecha_data[1:]:
            for col, value in enumerate(row_data):
                cell = tobacco_ws.cell(row=start_row, column=col+1)
                cell.value = value
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Highlight pecha columns
                if col in [3, 4, 5]:  # Days 15-31, Pecha %, Pecha Count
                    cell.fill = pecha_fill

            start_row += 1

        # Update Laundry Analytics sheet
        laundry_ws = wb['Laundry Analytics']

        # Calculate absolute values for laundry
        laundry_pecha_data = [
            ['Brand', 'Product', 'Total Purchases', 'Days 1-14', 'Days 15-31 (Pecha)', 'Pecha %', 'Pecha Count'],
            ['Surf', 'Bar Kalamansi', 157, 92, 65, '41.4%', 65],
            ['Tide', 'Detergent Bar', 128, 79, 49, '38.3%', 49],
            ['Ariel', 'Powder', 127, 88, 39, '30.7%', 39],
            ['Downy', 'Garden Bloom', 91, 61, 30, '33.0%', 30],
            ['Surf', 'Powder Pack', 36, 26, 10, '27.8%', 10]
        ]

        # Insert in laundry sheet around row 30
        start_row = 30

        laundry_ws.merge_cells(f'A{start_row}:G{start_row}')
        laundry_ws[f'A{start_row}'] = 'PECHA DE PELIGRO - ABSOLUTE VALUES & PERCENTAGES'
        laundry_ws[f'A{start_row}'].font = section_font
        laundry_ws[f'A{start_row}'].fill = pecha_fill
        laundry_ws[f'A{start_row}'].alignment = left_align
        laundry_ws[f'A{start_row}'].border = border

        # Add headers
        start_row += 2
        for col, header in enumerate(laundry_pecha_data[0]):
            cell = laundry_ws.cell(row=start_row, column=col+1)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Add data
        start_row += 1
        for row_data in laundry_pecha_data[1:]:
            for col, value in enumerate(row_data):
                cell = laundry_ws.cell(row=start_row, column=col+1)
                cell.value = value
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Highlight pecha columns
                if col in [4, 5, 6]:  # Days 15-31, Pecha %, Pecha Count
                    cell.fill = pecha_fill

            start_row += 1

        # Add summary insights sheet
        insights_ws = wb.create_sheet("Pecha de Peligro Insights")

        # Title
        insights_ws.merge_cells('A1:F1')
        insights_ws['A1'] = 'PECHA DE PELIGRO ANALYSIS - KEY INSIGHTS'
        insights_ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2E5594')
        insights_ws['A1'].fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        insights_ws['A1'].alignment = center_align
        insights_ws['A1'].border = border

        # Definition section
        row = 3
        insights_ws.merge_cells(f'A{row}:F{row}')
        insights_ws[f'A{row}'] = '📅 WHAT IS PECHA DE PELIGRO?'
        insights_ws[f'A{row}'].font = section_font
        insights_ws[f'A{row}'].fill = pecha_fill
        insights_ws[f'A{row}'].alignment = left_align
        insights_ws[f'A{row}'].border = border

        definition_text = [
            '• Filipino term meaning "Date of Danger"',
            '• Days 15-31 of the month when money runs low before next payday',
            '• Most Filipinos get paid monthly (15th or 30th/31st)',
            '• Consumer behavior shifts to budget-conscious purchases'
        ]

        row += 1
        for text in definition_text:
            insights_ws[f'A{row}'] = text
            insights_ws[f'A{row}'].font = data_font
            insights_ws[f'A{row}'].alignment = left_align
            insights_ws[f'A{row}'].border = border
            insights_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Tobacco insights
        row += 1
        insights_ws.merge_cells(f'A{row}:F{row}')
        insights_ws[f'A{row}'] = '🚬 TOBACCO PECHA DE PELIGRO PATTERNS'
        insights_ws[f'A{row}'].font = section_font
        insights_ws[f'A{row}'].fill = pecha_fill
        insights_ws[f'A{row}'].alignment = left_align
        insights_ws[f'A{row}'].border = border

        tobacco_insights = [
            '• Marca Leon (63.6% - 7 out of 11 purchases): Strongest pecha effect, cheapest option',
            '• TM (38.1% - 8 out of 21 purchases): Budget brand shows high late-month demand',
            '• Marlboro (35.2% - 38 out of 108 purchases): Premium brand, moderate pecha effect',
            '• Camel (34.6% - 18 out of 52 purchases): Expensive brand, lower pecha impact',
            '• Pattern: Cheaper tobacco brands = higher pecha de peligro sales'
        ]

        row += 1
        for insight in tobacco_insights:
            insights_ws[f'A{row}'] = insight
            insights_ws[f'A{row}'].font = data_font
            insights_ws[f'A{row}'].alignment = left_align
            insights_ws[f'A{row}'].border = border
            insights_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Laundry insights
        row += 1
        insights_ws.merge_cells(f'A{row}:F{row}')
        insights_ws[f'A{row}'] = '🧺 LAUNDRY PECHA DE PELIGRO PATTERNS'
        insights_ws[f'A{row}'].font = section_font
        insights_ws[f'A{row}'].fill = pecha_fill
        insights_ws[f'A{row}'].alignment = left_align
        insights_ws[f'A{row}'].border = border

        laundry_insights = [
            '• Surf Bar (41.4% - 65 out of 157 purchases): Highest pecha effect, essential item',
            '• Tide Bar (38.3% - 49 out of 128 purchases): Strong late-month demand',
            '• Downy (33.0% - 30 out of 91 purchases): Fabric conditioner, consistent need',
            '• Ariel Powder (30.7% - 39 out of 127 purchases): Premium product, lower pecha impact',
            '• Surf Powder (27.8% - 10 out of 36 purchases): Bulk purchase, bought early month',
            '• Pattern: Bar soaps show higher pecha sales than powder detergents'
        ]

        row += 1
        for insight in laundry_insights:
            insights_ws[f'A{row}'] = insight
            insights_ws[f'A{row}'].font = data_font
            insights_ws[f'A{row}'].alignment = left_align
            insights_ws[f'A{row}'].border = border
            insights_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Business implications
        row += 1
        insights_ws.merge_cells(f'A{row}:F{row}')
        insights_ws[f'A{row}'] = '💡 BUSINESS IMPLICATIONS'
        insights_ws[f'A{row}'].font = section_font
        insights_ws[f'A{row}'].fill = pecha_fill
        insights_ws[f'A{row}'].alignment = left_align
        insights_ws[f'A{row}'].border = border

        business_insights = [
            '• Inventory: Stock more budget brands during days 15-31',
            '• Promotions: Target premium brands for days 1-14',
            '• Pricing: Consider pecha-friendly pricing for essentials',
            '• Marketing: Emphasize value and affordability in late month campaigns',
            '• Product Mix: Balance premium and budget options based on calendar'
        ]

        row += 1
        for insight in business_insights:
            insights_ws[f'A{row}'] = insight
            insights_ws[f'A{row}'].font = data_font
            insights_ws[f'A{row}'].alignment = left_align
            insights_ws[f'A{row}'].border = border
            insights_ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Adjust column widths
        for i in range(6):
            insights_ws.column_dimensions[chr(65 + i)].width = 20

        # Save the workbook
        wb.save(file_path)

        print(f"✅ Absolute Pecha de Peligro values added successfully!")
        print(f"📁 Updated file: {file_path}")
        print(f"📊 Added sections:")
        print(f"   • Tobacco absolute values: Shows actual transaction counts")
        print(f"   • Laundry absolute values: Days 1-14 vs Days 15-31")
        print(f"   • New insights worksheet: Complete pecha analysis")
        print(f"   • Business implications: Actionable recommendations")

        # Print summary of absolute values
        print(f"\n🚬 TOBACCO PECHA DE PELIGRO ABSOLUTE VALUES:")
        print(f"   • Marlboro: 38 out of 108 transactions (35.2%)")
        print(f"   • Camel: 18 out of 52 transactions (34.6%)")
        print(f"   • TM: 8 out of 21 transactions (38.1%)")
        print(f"   • Marca Leon: 7 out of 11 transactions (63.6%) - HIGHEST")

        print(f"\n🧺 LAUNDRY PECHA DE PELIGRO ABSOLUTE VALUES:")
        print(f"   • Surf Bar: 65 out of 157 transactions (41.4%) - HIGHEST")
        print(f"   • Tide Bar: 49 out of 128 transactions (38.3%)")
        print(f"   • Ariel Powder: 39 out of 127 transactions (30.7%)")
        print(f"   • Downy: 30 out of 91 transactions (33.0%)")

    except Exception as e:
        print(f"❌ Error adding absolute values: {str(e)}")

if __name__ == "__main__":
    add_absolute_pecha_values()